# -*- coding: utf-8 -*-
"""
黑珍珠 + 米其林简易 RAG 问答系统

这个文件给黑客松演示用，尽量写成“小白也能看懂”的结构：
1. 文本读取：PDF / CSV / TXT
2. 文本切块：把长文档切成小段
3. 向量化：TfidfVectorizer.fit / transform / fit_transform
4. 向量检索：FAISS IndexFlatL2 / IndexIVFFlat + add / search
5. 文本生成：Transformers from_pretrained / encode / generate / decode
6. 多轮对话：ConversationMemory 保存上下文
7. 异步处理：asyncio.gather / get_event_loop / run_until_complete / aiohttp client.get
8. 网页清洗：requests.Session / requests.get / BeautifulSoup.find_all / decompose

注意：Transformers 本地模型不是必须的。没有 model_name 时会使用“基于检索片段的规则生成”，
这样项目不会因为模型没下载而跑不起来。
"""

from __future__ import annotations

import asyncio
import csv
import os
import re
import time
import unicodedata
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor
from dataclasses import asdict, dataclass, field
from multiprocessing import Pool, Process
from typing import Any, Callable, Iterable

import numpy as np
import requests
from bs4 import BeautifulSoup
from pypdf import PdfReader
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.pipeline import Pipeline as SklearnPipeline
from sklearn.preprocessing import normalize as sk_normalize

try:
    import aiohttp
except Exception:  # pragma: no cover - optional in some local envs
    aiohttp = None

try:
    import faiss
except Exception:  # pragma: no cover
    faiss = None

AutoModelForCausalLM = None
AutoTokenizer = None
hf_pipeline = None


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(BASE_DIR)
BLACK_PEARL_PDF_PATH = os.environ.get(
    "BLACK_PEARL_PDF_PATH",
    os.path.join(PROJECT_DIR, "黑珍珠-米其林.pdf"),
)
MICHELIN_CSV_PATH = os.environ.get(
    "MICHELIN_CSV_PATH",
    os.path.join(BASE_DIR, "rag_documents.csv"),
)
BLACK_PEARL_SINGAPORE_XLSX_PATH = os.environ.get(
    "BLACK_PEARL_SINGAPORE_XLSX_PATH",
    os.path.join(PROJECT_DIR, "黑珍珠-rag数据-新加坡.xlsx"),
)

_CHAR_MAP = str.maketrans({
    "⿊": "黑",
    "⽶": "米",
    "⻓": "长",
    "⻘": "青",
    "⻔": "门",
    "⻥": "鱼",
    "⻋": "车",
    "⻝": "食",
    "⻰": "龙",
    "⻩": "黄",
    "⺠": "民",
})
CITY_ALIASES = {
    "北京": ["北京", "Beijing"],
    "上海": ["上海", "Shanghai"],
    "广州": ["广州", "Guangzhou"],
    "深圳": ["深圳", "Shenzhen"],
    "杭州": ["杭州", "Hangzhou"],
    "成都": ["成都", "Chengdu"],
    "香港": ["香港", "Hong Kong"],
    "澳门": ["澳门", "Macau", "Macao"],
    "台北": ["台北", "Taipei"],
    "新加坡": ["新加坡", "Singapore"],
}
RAG_LATENCY_BUDGET_SECONDS = float(os.environ.get("RAG_LATENCY_BUDGET_SECONDS", "10"))
QUERY_CACHE_TTL_SECONDS = int(os.environ.get("RAG_QUERY_CACHE_TTL_SECONDS", "180"))
QUERY_CACHE_MAX_SIZE = int(os.environ.get("RAG_QUERY_CACHE_MAX_SIZE", "128"))
_TEXT_CACHE: dict[tuple[str, float], str] = {}
_DOC_CACHE: dict[tuple[Any, ...], list["RAGDocument"]] = {}
_QUERY_CACHE: "OrderedDict[tuple[Any, ...], tuple[float, dict[str, Any]]]" = OrderedDict()


@dataclass
class RAGDocument:
    text: str
    source: str
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass
class SearchHit:
    text: str
    source: str
    score: float
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass
class RAGAnswer:
    answer: str
    query: str
    hits: list[SearchHit]
    history_used: list[dict[str, str]]
    structured: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["references"] = sorted({h.source for h in self.hits})
        return data


class ConversationMemory:
    """保存多轮上下文，让下一轮问题能参考上一轮问答。"""

    def __init__(self, max_turns: int = 6):
        self.max_turns = max_turns
        self.messages: list[dict[str, str]] = []

    def add_user(self, content: str) -> None:
        self.messages.append({"role": "user", "content": str(content or "")})
        self._trim()

    def add_assistant(self, content: str) -> None:
        self.messages.append({"role": "assistant", "content": str(content or "")})
        self._trim()

    def context_text(self) -> str:
        return "\n".join(f"{m['role']}: {m['content']}" for m in self.messages[-self.max_turns * 2:])

    def snapshot(self) -> list[dict[str, str]]:
        return list(self.messages[-self.max_turns * 2:])

    def _trim(self) -> None:
        keep = self.max_turns * 2
        if len(self.messages) > keep:
            self.messages = self.messages[-keep:]


def normalize_text(text: str) -> str:
    text = unicodedata.normalize("NFKC", str(text or "")).translate(_CHAR_MAP)
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]", " ", text)
    return re.sub(r"[ \t]+", " ", text).strip()


def contains_city_alias(text: str, aliases: list[str]) -> bool:
    raw = str(text or "")
    raw_lower = raw.lower()
    for alias in aliases:
        if re.search(r"[A-Za-z]", alias):
            if re.search(rf"\b{re.escape(alias.lower())}\b", raw_lower):
                return True
        elif alias in raw:
            return True
    return False


def read_pdf_text(path: str = BLACK_PEARL_PDF_PATH) -> str:
    mtime = os.path.getmtime(path)
    cache_key = ("pdf", path, mtime)
    if cache_key in _TEXT_CACHE:
        return _TEXT_CACHE[cache_key]
    reader = PdfReader(path)
    text = "\n".join((page.extract_text() or "") for page in reader.pages)
    text = normalize_text(text)
    _TEXT_CACHE[cache_key] = text
    return text


def read_michelin_csv(path: str = MICHELIN_CSV_PATH, limit: int = 260) -> str:
    if not os.path.exists(path):
        return ""
    mtime = os.path.getmtime(path)
    cache_key = ("csv_head", path, mtime, limit)
    if cache_key in _TEXT_CACHE:
        return _TEXT_CACHE[cache_key]
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for idx, row in enumerate(reader):
            if idx >= limit:
                break
            rows.append(" | ".join(str(v or "") for v in row.values()))
    text = normalize_text("\n".join(rows))
    _TEXT_CACHE[cache_key] = text
    return text


def read_michelin_csv_for_query(query: str, path: str = MICHELIN_CSV_PATH, limit: int = 360) -> str:
    if not os.path.exists(path):
        return ""
    mtime = os.path.getmtime(path)
    normalized_query = normalize_text(query).lower()
    cache_key = ("csv_query", path, mtime, normalized_query, limit)
    if cache_key in _TEXT_CACHE:
        return _TEXT_CACHE[cache_key]
    city_terms = re.findall(r"(北京|上海|广州|深圳|杭州|成都|重庆|西安|武汉|南京|苏州|长沙|青岛|厦门|天津|三亚|香港|澳门|台北|新加坡)", query)
    city_aliases = []
    for city in city_terms:
        city_aliases.extend(CITY_ALIASES.get(city, [city]))
    stop_terms = {"米其林", "餐厅", "推荐", "哪些", "哪里", "一下", *city_terms, *city_aliases}
    query_for_terms = normalize_text(query)
    for term in sorted(stop_terms, key=len, reverse=True):
        query_for_terms = query_for_terms.replace(term, " ")
    query_terms = [t for t in re.findall(r"[\u4e00-\u9fff]{2,}|[A-Za-z]{3,}", query_for_terms) if t not in stop_terms]
    matched, fallback = [], []
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            raw_text = " | ".join(str(v or "") for v in row.values())
            city_scope = " | ".join(str(row.get(k) or "") for k in ("location", "country"))
            raw_lower = raw_text.lower()
            if len(fallback) < 80:
                fallback.append(normalize_text(raw_text))
            city_ok = not city_aliases or contains_city_alias(city_scope, city_aliases)
            term_ok = not query_terms or any(term.lower() in raw_lower for term in query_terms)
            if city_ok and term_ok:
                matched.append(normalize_text(raw_text))
            if len(matched) >= limit:
                break
    text = "\n".join(matched or fallback)
    _TEXT_CACHE[cache_key] = text
    return text


def read_black_pearl_singapore_xlsx(path: str = BLACK_PEARL_SINGAPORE_XLSX_PATH) -> str:
    """读取新加坡黑珍珠 Excel，并结构化成 RAG 文本。"""
    if not os.path.exists(path):
        return ""
    mtime = os.path.getmtime(path)
    cache_key = ("black_pearl_singapore_xlsx", path, mtime)
    if cache_key in _TEXT_CACHE:
        return _TEXT_CACHE[cache_key]

    rows = []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        values = list(ws.iter_rows(values_only=True))
        if not values:
            return ""
        headers = [str(h or "").strip() for h in values[0]]
        for raw in values[1:]:
            item = {headers[i]: raw[i] if i < len(raw) else "" for i in range(len(headers))}
            name = str(item.get("餐厅名称") or "").strip()
            if not name:
                continue
            cuisine = str(item.get("菜系") or "").strip()
            diamond = str(item.get("钻级") or "").strip()
            price = str(item.get("人均消费（元）") or "").strip()
            place = str(item.get("地点") or "新加坡").strip()
            rows.append(
                f"{place}-黑珍珠新加坡 | 餐厅名称:{name} | 菜系:{cuisine} | 钻级:{diamond} | 人均消费约¥{price} | 地点:{place}"
            )
    except Exception:
        import pandas as pd
        df = pd.read_excel(path)
        for _, item in df.iterrows():
            name = str(item.get("餐厅名称") or "").strip()
            if not name or name == "nan":
                continue
            cuisine = str(item.get("菜系") or "").strip()
            diamond = str(item.get("钻级") or "").strip()
            price = str(item.get("人均消费（元）") or "").strip()
            place = str(item.get("地点") or "新加坡").strip()
            rows.append(
                f"{place}-黑珍珠新加坡 | 餐厅名称:{name} | 菜系:{cuisine} | 钻级:{diamond} | 人均消费约¥{price} | 地点:{place}"
            )

    text = normalize_text("\n".join(rows))
    _TEXT_CACHE[cache_key] = text
    return text


def split_text(text: str, chunk_size: int = 760, overlap: int = 120) -> list[str]:
    text = normalize_text(text)
    section_parts = [
        p.strip()
        for p in re.split(r"\n{2,}|(?=^[\u2e80-\u9fff]{1,10}\s*[-—－]\s*(?:黑珍珠|米其林)[:：])", text, flags=re.M)
        if len(p.strip()) >= 40
    ]
    if len(section_parts) >= 5:
        return section_parts
    chunks = []
    step = max(80, chunk_size - overlap)
    flat = re.sub(r"\s+", " ", text)
    for start in range(0, len(flat), step):
        chunk = flat[start:start + chunk_size].strip()
        if len(chunk) >= 40:
            chunks.append(chunk)
    return chunks


def build_documents(
    black_pearl_pdf: str = BLACK_PEARL_PDF_PATH,
    michelin_csv: str = MICHELIN_CSV_PATH,
    singapore_xlsx: str = BLACK_PEARL_SINGAPORE_XLSX_PATH,
) -> list[RAGDocument]:
    pdf_mtime = os.path.getmtime(black_pearl_pdf) if black_pearl_pdf and os.path.exists(black_pearl_pdf) else 0
    csv_mtime = os.path.getmtime(michelin_csv) if michelin_csv and os.path.exists(michelin_csv) else 0
    xlsx_mtime = os.path.getmtime(singapore_xlsx) if singapore_xlsx and os.path.exists(singapore_xlsx) else 0
    cache_key = ("docs", black_pearl_pdf, pdf_mtime, michelin_csv, csv_mtime, singapore_xlsx, xlsx_mtime)
    if cache_key in _DOC_CACHE:
        return list(_DOC_CACHE[cache_key])

    def load_black_pearl_docs() -> list[RAGDocument]:
        if not black_pearl_pdf or not os.path.exists(black_pearl_pdf):
            return []
        return [
            RAGDocument(
                text=chunk,
                source=os.path.basename(black_pearl_pdf),
                metadata={"kind": "black_pearl_pdf", "chunk_id": idx},
            )
            for idx, chunk in enumerate(split_text(read_pdf_text(black_pearl_pdf)))
        ]

    def load_michelin_docs() -> list[RAGDocument]:
        michelin_text = read_michelin_csv(michelin_csv) if michelin_csv else ""
        if not michelin_text:
            return []
        return [
            RAGDocument(
                text=chunk,
                source=os.path.basename(michelin_csv),
                metadata={"kind": "michelin_csv", "chunk_id": idx},
            )
            for idx, chunk in enumerate(split_text(michelin_text, chunk_size=560, overlap=80))
        ]

    def load_singapore_docs() -> list[RAGDocument]:
        singapore_text = read_black_pearl_singapore_xlsx(singapore_xlsx) if singapore_xlsx else ""
        if not singapore_text:
            return []
        return [
            RAGDocument(
                text=line,
                source=os.path.basename(singapore_xlsx),
                metadata={"kind": "black_pearl_singapore_xlsx", "chunk_id": idx, "city": "新加坡"},
            )
            for idx, line in enumerate(singapore_text.splitlines())
            if line.strip()
        ]

    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = [
            executor.submit(load_black_pearl_docs),
            executor.submit(load_michelin_docs),
            executor.submit(load_singapore_docs),
        ]
        documents = []
        for future in futures:
            documents.extend(future.result())
    _DOC_CACHE[cache_key] = documents
    return list(documents)


def build_sklearn_pipeline() -> SklearnPipeline:
    """sklearn.pipeline.Pipeline：把预处理和 TF-IDF 放进一个管道。"""
    return SklearnPipeline([
        ("tfidf", TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 4), max_features=4096)),
    ])


def pipeline() -> SklearnPipeline:
    """pipeline()：给演示用的 sklearn.pipeline 入口。"""
    return build_sklearn_pipeline()


def IndexIVF(nlist: int, d: int, metric: int | None = None):
    """IndexIVF(nlist, d, metric)：用 IndexIVFFlat 创建倒排文件索引。"""
    if faiss is None:
        raise RuntimeError("faiss 未安装，无法创建 IndexIVF")
    metric = faiss.METRIC_L2 if metric is None else metric
    quantizer = faiss.IndexFlatL2(d)
    return faiss.IndexIVFFlat(quantizer, d, nlist, metric)


def IndexIVFFlat(nlist: int, d: int, metric: int | None = None):
    """IndexIVFFlat(nlist, d, metric)：FAISS 倒排索引便捷入口。"""
    return IndexIVF(nlist, d, metric)


class TransformerGenerator:
    """Transformers 文本生成模块。展示 temperature / top_k / top_p / generate 等参数。"""

    def __init__(self, model_name: str | None = None):
        self.model_name = model_name
        self.tokenizer = None
        self.model = None
        self.generator = None
        if model_name:
            global AutoModelForCausalLM, AutoTokenizer, hf_pipeline
            if AutoTokenizer is None or AutoModelForCausalLM is None:
                from transformers import AutoModelForCausalLM as _AutoModelForCausalLM
                from transformers import AutoTokenizer as _AutoTokenizer
                from transformers import pipeline as _hf_pipeline
                AutoTokenizer = _AutoTokenizer
                AutoModelForCausalLM = _AutoModelForCausalLM
                hf_pipeline = _hf_pipeline
            # from_pretrained(model_name)：加载指定预训练模型及参数
            self.tokenizer = AutoTokenizer.from_pretrained(model_name)
            self.model = AutoModelForCausalLM.from_pretrained(model_name)
            if hf_pipeline:
                self.generator = hf_pipeline("text-generation", model=self.model, tokenizer=self.tokenizer)

    def generate(
        self,
        prompt: str,
        max_length: int = 512,
        temperature: float = 0.2,
        top_k: int = 40,
        top_p: float = 0.9,
    ) -> str:
        if not self.model or not self.tokenizer:
            return fallback_generate(prompt)
        # tokenizer.encode(prompt, return_tensors="pt")：转换成 PyTorch 张量
        inputs = self.tokenizer.encode(prompt, return_tensors="pt")
        # generate(inputs, max_length)：temperature/top_k/top_p 控制生成随机性和候选范围
        tokens = self.model.generate(
            inputs,
            max_length=max_length,
            do_sample=True,
            temperature=temperature,
            top_k=top_k,
            top_p=top_p,
            pad_token_id=self.tokenizer.eos_token_id,
        )
        # tokenizer.decode(tokens)：解码成可读文本
        return self.tokenizer.decode(tokens[0], skip_special_tokens=True)


def fallback_generate(prompt: str) -> str:
    """没有本地生成模型时，用检索片段组织一段稳定回答。"""
    context = prompt.split("【检索片段】", 1)[-1].strip()
    context = context.split("\n\n请输出", 1)[0].strip()
    lines = [x.strip() for x in re.split(r"\n+", context) if x.strip()]
    picked = lines[:4]
    if not picked:
        return "没有检索到足够资料，请换一个城市、预算或菜系关键词。"
    return "根据本地 RAG 检索结果，可以这样回答：\n" + "\n".join(f"{i + 1}. {x[:260]}" for i, x in enumerate(picked))


def tune_generation_params(query: str, history: list[dict[str, str]] | None = None) -> dict[str, Any]:
    """根据问题动态调节检索和生成参数，兼顾速度和质量。"""
    q = normalize_text(query)
    has_city = bool(re.search(r"(北京|上海|广州|深圳|杭州|成都|重庆|西安|武汉|南京|苏州|长沙|青岛|厦门|天津|三亚|香港|澳门|台北|新加坡)", q))
    is_follow_up = bool(history) and bool(re.search(r"为什么|原因|那|继续|详细|适合|换|还有|对比|怎么选", q))
    is_broad = bool(re.search(r"为什么|详细|对比|分析|怎么选|适合", q)) or len(q) > 28
    top_k = 3 if has_city and not is_broad else 5
    if is_follow_up:
        top_k = max(top_k, 5)
    return {
        "top_k": top_k,
        "max_length": 420 if not is_broad else 560,
        "temperature": 0.15 if has_city else 0.25,
        "top_p": 0.85 if has_city else 0.92,
        "search_multiplier": 8 if has_city else 5,
        "candidate_cap": 96 if has_city else 48,
    }


class SimpleRAGSystem:
    """一个完整但轻量的 RAG：向量化 -> FAISS 检索 -> 生成 -> 多轮记忆。"""

    def __init__(self, model_name: str | None = None, use_ivf: bool = False, nlist: int = 32):
        self.documents: list[RAGDocument] = []
        self.vectorizer = TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 4), max_features=4096)
        self.pipeline = build_sklearn_pipeline()
        self.doc_vectors = None
        self.index = None
        self.use_ivf = use_ivf
        self.nlist = nlist
        self.generator = TransformerGenerator(model_name=model_name)
        self.memory = ConversationMemory()

    def fit(self, documents: list[RAGDocument]) -> "SimpleRAGSystem":
        """fit(documents)：训练 TF-IDF 向量化器并创建 FAISS 索引。"""
        self.documents = documents
        texts = [d.text for d in documents]
        # fit_transform(documents)：训练并转换文档向量
        sparse_vectors = self.vectorizer.fit_transform(texts)
        # np.array(doc_vectors)：转成 NumPy 数组供 FAISS 使用
        self.doc_vectors = np.array(sk_normalize(sparse_vectors).toarray(), dtype="float32")
        self._build_faiss_index(self.doc_vectors)
        return self

    def transform(self, query: str) -> np.ndarray:
        """transform(query)：把查询转成 TF-IDF 向量。"""
        sparse_query = self.vectorizer.transform([query])
        return np.array(sk_normalize(sparse_query).toarray(), dtype="float32")

    def _build_faiss_index(self, vectors: np.ndarray) -> None:
        if vectors.size == 0:
            self.index = None
            return
        dimension = vectors.shape[1]
        if faiss is None:
            self.index = None
            return
        if self.use_ivf and len(vectors) >= max(self.nlist * 40, 128):
            # IndexIVFFlat(nlist, d, metric)：倒排索引，适合较大文档库快速检索
            quantizer = faiss.IndexFlatL2(dimension)
            index = faiss.IndexIVFFlat(quantizer, dimension, self.nlist, faiss.METRIC_L2)
            index.train(vectors)
            # add(vectors)：把文档向量加入 FAISS 索引
            index.add(vectors)
            self.index = index
            return
        # IndexFlatL2(dimension)：小数据量直接用 L2 暴力检索
        index = faiss.IndexFlatL2(dimension)
        index.add(vectors)
        self.index = index

    def retrieve(self, query: str, top_k: int = 4) -> list[SearchHit]:
        """search(query_vector, top_k)：检索最相似的 top_k 文档。"""
        if not self.documents:
            return []
        query_vector = self.transform(query)
        if self.index is not None:
            candidates = []
            cities = re.findall(r"(北京|上海|广州|深圳|杭州|成都|重庆|西安|武汉|南京|苏州|长沙|青岛|厦门|天津|三亚|香港|澳门|新加坡)", query)
            city_aliases = {city: CITY_ALIASES.get(city, [city]) for city in cities}
            wants_michelin = "米其林" in query or "michelin" in query.lower()
            wants_black_pearl = "黑珍珠" in query
            cuisine_terms = [t for t in ("中餐", "法国菜", "日本菜", "英国菜", "娘惹菜", "海鲜", "西餐", "粤菜", "川菜") if t in query]
            tuned = tune_generation_params(query)
            search_k = min(
                len(self.documents),
                max(top_k * tuned["search_multiplier"], top_k),
                tuned["candidate_cap"],
            )
            distances, indices = self.index.search(query_vector, search_k)
            seen_indices = set()
            for distance, idx in zip(distances[0], indices[0]):
                if idx < 0:
                    continue
                seen_indices.add(int(idx))
                doc = self.documents[int(idx)]
                city_boost = 0.0
                for city in cities:
                    if re.search(rf"^\s*{re.escape(city)}\s*[-—－]\s*(?:黑珍珠|米其林)", doc.text):
                        city_boost += 20.0
                    elif contains_city_alias(doc.text, city_aliases.get(city, [city])):
                        city_boost += 0.5
                source_boost = 0.0
                kind = str(doc.metadata.get("kind", ""))
                if wants_michelin and kind == "michelin_csv":
                    source_boost += 30.0
                if wants_michelin and kind == "black_pearl_pdf" and not wants_black_pearl:
                    source_boost -= 12.0
                if wants_black_pearl and kind == "black_pearl_singapore_xlsx" and any(c == "新加坡" for c in cities):
                    source_boost += 80.0
                elif wants_black_pearl and kind in ("black_pearl_pdf", "black_pearl_singapore_xlsx"):
                    source_boost += 30.0
                term_boost = sum(18.0 for term in cuisine_terms if term in doc.text)
                score = float(-distance + city_boost + source_boost + term_boost)
                candidates.append(SearchHit(doc.text, doc.source, score, doc.metadata))
            if cities:
                for idx, doc in enumerate(self.documents):
                    if idx in seen_indices:
                        continue
                    if not any(contains_city_alias(doc.text, city_aliases.get(city, [city])) for city in cities):
                        continue
                    kind = str(doc.metadata.get("kind", ""))
                    source_boost = 80.0 if (wants_black_pearl and kind == "black_pearl_singapore_xlsx" and any(c == "新加坡" for c in cities)) else (30.0 if (wants_michelin and kind == "michelin_csv") or (wants_black_pearl and kind in ("black_pearl_pdf", "black_pearl_singapore_xlsx")) else 0.0)
                    term_boost = sum(18.0 for term in cuisine_terms if term in doc.text)
                    candidates.append(SearchHit(doc.text, doc.source, 12.0 + source_boost + term_boost, doc.metadata))
            return sorted(candidates, key=lambda h: h.score, reverse=True)[:top_k]
        scores = self.doc_vectors @ query_vector[0]
        best = np.argsort(scores)[::-1][:top_k]
        return [SearchHit(self.documents[i].text, self.documents[i].source, float(scores[i]), self.documents[i].metadata) for i in best]

    def ask(
        self,
        query: str,
        top_k: int = 4,
        max_length: int = 512,
        temperature: float = 0.2,
        top_p: float = 0.9,
    ) -> RAGAnswer:
        self.memory.add_user(query)
        query_with_context = f"{self.memory.context_text()}\n当前问题: {query}"
        hits = self.retrieve(query_with_context, top_k=top_k)
        context = "\n\n".join(f"[{i + 1}] {h.text}" for i, h in enumerate(hits))
        prompt = (
            "你是米其林和黑珍珠餐厅 RAG 助手。只能依据检索片段回答，资料不足要说明。\n"
            f"【多轮上下文】\n{self.memory.context_text()}\n\n"
            f"【用户问题】\n{query}\n\n"
            f"【检索片段】\n{context}\n\n"
            "请输出：推荐餐厅/城市/理由/人均或星钻信息/注意事项。"
        )
        answer = self.generator.generate(
            prompt,
            max_length=max_length,
            temperature=temperature,
            top_k=top_k,
            top_p=top_p,
        )
        self.memory.add_assistant(answer)
        return RAGAnswer(
            answer=answer,
            query=query,
            hits=hits,
            history_used=self.memory.snapshot(),
            structured={
                "top_k": top_k,
                "temperature": temperature,
                "top_p": top_p,
                "vectorizer": "TfidfVectorizer",
                "index": type(self.index).__name__ if self.index is not None else "numpy_dot_fallback",
            },
        )


_RAG_CACHE: dict[str, Any] = {"mtime": None, "rag": None}


def _history_signature(history: list[dict[str, str]] | None) -> tuple[str, ...]:
    if not history:
        return ()
    return tuple(f"{m.get('role','')}:{normalize_text(m.get('content',''))[:120]}" for m in history[-6:])


def _data_signature() -> tuple[float, float, float]:
    return (
        os.path.getmtime(BLACK_PEARL_PDF_PATH) if os.path.exists(BLACK_PEARL_PDF_PATH) else 0,
        os.path.getmtime(MICHELIN_CSV_PATH) if os.path.exists(MICHELIN_CSV_PATH) else 0,
        os.path.getmtime(BLACK_PEARL_SINGAPORE_XLSX_PATH) if os.path.exists(BLACK_PEARL_SINGAPORE_XLSX_PATH) else 0,
    )


def _get_cached_query(cache_key: tuple[Any, ...]) -> dict[str, Any] | None:
    item = _QUERY_CACHE.get(cache_key)
    if not item:
        return None
    ts, data = item
    if time.time() - ts > QUERY_CACHE_TTL_SECONDS:
        _QUERY_CACHE.pop(cache_key, None)
        return None
    _QUERY_CACHE.move_to_end(cache_key)
    cached = dict(data)
    cached.setdefault("structured", {})
    cached["structured"] = dict(cached["structured"])
    cached["structured"]["cache_hit"] = True
    return cached


def _set_cached_query(cache_key: tuple[Any, ...], data: dict[str, Any]) -> None:
    _QUERY_CACHE[cache_key] = (time.time(), data)
    _QUERY_CACHE.move_to_end(cache_key)
    while len(_QUERY_CACHE) > QUERY_CACHE_MAX_SIZE:
        _QUERY_CACHE.popitem(last=False)


def get_default_rag(force_reload: bool = False, model_name: str | None = None) -> SimpleRAGSystem:
    paths = [BLACK_PEARL_PDF_PATH, MICHELIN_CSV_PATH, BLACK_PEARL_SINGAPORE_XLSX_PATH]
    mtimes = tuple(os.path.getmtime(p) if os.path.exists(p) else 0 for p in paths)
    if not force_reload and _RAG_CACHE["rag"] is not None and _RAG_CACHE["mtime"] == mtimes:
        return _RAG_CACHE["rag"]
    documents = build_documents()
    rag = SimpleRAGSystem(model_name=model_name)
    rag.fit(documents)
    _RAG_CACHE.update({"mtime": mtimes, "rag": rag})
    return rag


def query_michelin_black_pearl(
    query: str,
    history: list[dict[str, str]] | None = None,
    top_k: int | None = None,
    model_name: str | None = None,
) -> dict[str, Any]:
    started = time.perf_counter()
    tuned = tune_generation_params(query, history)
    effective_top_k = int(top_k or tuned["top_k"])
    mode = "michelin" if ("米其林" in query or "michelin" in query.lower()) else "mixed"
    cache_key = (
        "qa",
        normalize_text(query).lower(),
        _history_signature(history),
        effective_top_k,
        mode,
        model_name or "",
        _data_signature(),
    )
    cached = _get_cached_query(cache_key)
    if cached:
        return cached

    wants_michelin = "米其林" in query or "michelin" in query.lower()
    wants_black_pearl = "黑珍珠" in query
    if wants_michelin:
        documents = build_documents(michelin_csv="") if wants_black_pearl else []
        michelin_text = read_michelin_csv_for_query(query)
        if michelin_text:
            for idx, chunk in enumerate(split_text(michelin_text, chunk_size=560, overlap=80)):
                documents.append(RAGDocument(
                    text=chunk,
                    source=os.path.basename(MICHELIN_CSV_PATH),
                    metadata={"kind": "michelin_csv", "chunk_id": idx, "filtered_by_query": True},
                ))
        rag = SimpleRAGSystem(model_name=model_name)
        rag.fit(documents)
    else:
        rag = get_default_rag(model_name=model_name)
    if history:
        rag.memory.messages = history[-12:]
    result = rag.ask(
        query,
        top_k=effective_top_k,
        max_length=tuned["max_length"],
        temperature=tuned["temperature"],
        top_p=tuned["top_p"],
    )
    data = result.to_dict()
    data.setdefault("structured", {})
    data["structured"].update({
        "cache_hit": False,
        "latency_ms": round((time.perf_counter() - started) * 1000),
        "latency_budget_seconds": RAG_LATENCY_BUDGET_SECONDS,
        "dynamic_params": tuned,
    })
    _set_cached_query(cache_key, data)
    return data


def query_black_pearl_pdf(query: str, pdf_path: str = BLACK_PEARL_PDF_PATH, top_k: int = 4) -> dict[str, Any]:
    documents = [
        RAGDocument(text=chunk, source=os.path.basename(pdf_path), metadata={"kind": "black_pearl_pdf", "chunk_id": idx})
        for idx, chunk in enumerate(split_text(read_pdf_text(pdf_path)))
    ]
    rag = SimpleRAGSystem(model_name=None)
    rag.fit(documents)
    return rag.ask(query, top_k=top_k).to_dict()


# multiprocessing.Pool.map(func, inputs)：并行处理多个输入
def parallel_map(func: Callable[[Any], Any], inputs: Iterable[Any]) -> list[Any]:
    with Pool() as pool:
        return pool.map(func, inputs)


# Process.join()：阻塞主进程，等待子进程结束
def process_join_demo(target: Callable[..., Any], args: tuple[Any, ...] = ()) -> int | None:
    process = Process(target=target, args=args)
    process.start()
    process.join()
    return process.exitcode


def clean_html(html: str, tag: str = "p", attrs: dict[str, Any] | None = None) -> list[str]:
    soup = BeautifulSoup(html or "", "html.parser")
    for bad in soup.find_all(["script", "style", "noscript"]):
        bad.decompose()
    nodes = soup.find_all(tag, attrs or {})
    return [normalize_text(n.get_text(" ")) for n in nodes if normalize_text(n.get_text(" "))]


def fetch_url_sync(url: str) -> list[str]:
    with requests.Session() as session:
        response = session.get(url, timeout=8)
        response.raise_for_status()
        return clean_html(response.text)


def fetch_url_once(url: str) -> str:
    response = requests.get(url, timeout=8)
    response.raise_for_status()
    return response.text


async def fetch_url_async(url: str) -> list[str]:
    if aiohttp is None:
        return fetch_url_sync(url)
    async with aiohttp.ClientSession() as client:
        async with client.get(url, timeout=8) as response:
            html = await response.text()
            return clean_html(html)


async def fetch_many_async(urls: list[str]) -> list[list[str]]:
    tasks = [fetch_url_async(url) for url in urls]
    return await asyncio.gather(*tasks)


def run_async_fetch(urls: list[str]) -> list[list[str]]:
    try:
        loop = asyncio.get_event_loop()
    except RuntimeError:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
    return loop.run_until_complete(fetch_many_async(urls))


# 兼容旧代码命名：load_black_pearl_data / create_vector_db / query_with_deepseek
def load_black_pearl_data() -> list[RAGDocument]:
    return [
        RAGDocument(text=chunk, source=os.path.basename(BLACK_PEARL_PDF_PATH), metadata={"kind": "black_pearl_pdf"})
        for chunk in split_text(read_pdf_text(BLACK_PEARL_PDF_PATH))
    ]


def create_vector_db() -> SimpleRAGSystem:
    rag = SimpleRAGSystem(model_name=None)
    rag.fit(build_documents())
    return rag


def query_with_deepseek(query: str, context: str) -> str:
    prompt = f"【用户问题】{query}\n【检索片段】\n{context}"
    return fallback_generate(prompt)


def main() -> None:
    try:
        import streamlit as st
    except Exception:
        print("Streamlit 未安装。可直接运行：python hei_zhen_zhu_local.py '北京黑珍珠人均500餐厅'")
        return

    st.set_page_config(page_title="黑珍珠/米其林 RAG", page_icon="💎", layout="centered")
    st.title("💎 黑珍珠/米其林简易 RAG 问答")
    st.caption("PDF/CSV -> TF-IDF -> FAISS -> 检索片段 -> 生成回答 -> 多轮上下文")
    if "rag" not in st.session_state:
        with st.spinner("正在构建向量索引..."):
            st.session_state.rag = get_default_rag(force_reload=True)
    if "messages" not in st.session_state:
        st.session_state.messages = []
    for msg in st.session_state.messages:
        with st.chat_message(msg["role"]):
            st.write(msg["content"])
    if prompt := st.chat_input("例如：北京黑珍珠人均500左右有什么推荐？"):
        st.session_state.messages.append({"role": "user", "content": prompt})
        with st.chat_message("user"):
            st.write(prompt)
        with st.chat_message("assistant"):
            result = st.session_state.rag.ask(prompt)
            st.write(result.answer)
            st.caption("参考：" + "、".join(sorted({h.source for h in result.hits})))
        st.session_state.messages.append({"role": "assistant", "content": result.answer})


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1:
        question = " ".join(sys.argv[1:])
        t0 = time.perf_counter()
        result = query_michelin_black_pearl(question)
        print(f"输出内容: {result['answer']}")
        print(f"参考来源: {result['references']}")
        print(f"耗时: {round((time.perf_counter() - t0) * 1000)}ms")
    else:
        main()
